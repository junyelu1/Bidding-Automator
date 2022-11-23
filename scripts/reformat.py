import pandas as pd
import numpy as np
from openpyxl import *
from openpyxl.styles import Border, Side, Alignment, colors, Font
import re


def bidExcelFormat(filePath: str):

    assert ".xlsx" in filePath, f"Input FilePath given is not Excel"

    # Using OpenPyXL to finish formating and grouping
    wb = load_workbook(filePath)

    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        rowcount = sheet.max_row
        colcount = sheet.max_column

        # Default Border, Alignment
        border = Border(left=Side(border_style='thin', color=colors.BLACK), right=Side(border_style='thin', color=colors.BLACK),
                        top=Side(border_style='thin', color=colors.BLACK), bottom=Side(border_style='thin', color=colors.BLACK))
        leftalign = Alignment(
            horizontal='left', vertical='center', wrap_text=True)
        centeralign = Alignment(
            horizontal='center', vertical='center', wrap_text=True)

        # Find the number of unique Baos
        bao = [cell.value for cell in sheet['B']]
        baoNum = len(set(bao))

        rowsToBeAdded = [2]
        i = 2

        # Set Up Subtotal Structures used in Excel
        while(i < rowcount+baoNum+1):

            if sheet.cell(i, 2).value != sheet.cell(i+1, 2).value:
                rowsToBeAdded.append(i)
                sheet.insert_rows(i+1)
                sheet.cell(i+1, 2).value = sheet.cell(i, 2).value+" Total"
                sheet.cell(i+1, 2).font = Font(bold=True)
                sheet.cell(i+1, 14).value = "=SUBTOTAL(9,N" + \
                    str(rowsToBeAdded[-2])+":N"+str(i)+")"
                sheet.cell(i+1, 16).value = "=SUBTOTAL(9,P" + \
                    str(rowsToBeAdded[-2])+":P"+str(i)+")"
                sheet.cell(i+1, 17).value = "=1-P"+str(i)+"/N"+str(i)
                i += 1
            i += 1

        # Update Rowcount and setup Grand Total
        rowcount = sheet.max_row
        sheet.cell(rowcount-1, 2).value = "Grand Total"
        sheet.cell(rowcount-1, 2).font = Font(bold=True)
        sheet.cell(rowcount-1, 14).value = "=SUBTOTAL(9, N2:N" + \
            str(rowcount-3)+")"
        sheet.cell(rowcount-1, 14).value = "=SUBTOTAL(9, P2:P" + \
            str(rowcount-3)+")"

       # Format columns using predefined rules, using col numbers instead for simplicity
        for i in range(2, rowcount):
            for j in range(1, colcount+1):
                sheet.cell(i, j).border = border
                if j in [4, 5, 8]:
                    sheet.cell(i, j).alignment = leftalign
                else:
                    if j == 17:
                        sheet.cell(i, j).number_format = "0.00%"
                    sheet.cell(i, j).alignment = centeralign

        # Column Width Adjustments, find column width
        for col in sheet.columns:
            curwidth = len(re.sub(
                "[A-Za-z0-9\!\%\[\]\,\。]", "", str(col[0].value))) + len(str(col[0].value))
            for j in range(len(col)):
                chinese = len(
                    re.sub("[A-Za-z0-9\!\%\[\]\,\。]", "", str(col[j].value)))
                if curwidth < chinese + len(str(col[j].value)):
                    curwidth = chinese + len(str(col[j].value))
            sheet.column_dimensions[col[0].column_letter].width = curwidth+0.8

        # Row Width
        for i in range(1, rowcount):
            sheet.row_dimensions[i].height = 17

        # Hide Irrelavent columns
        clients = [cell.value for cell in sheet['E']]
        invisible = ['A', 'C', 'F', 'G', 'I',
                     'R', 'U', 'V', 'X', 'Y', 'Z', 'AA']
        if len(set(clients)) > 2:
            invisible = invisible + ['D']
        else:
            invisible = invisible + ['E']
        for col in invisible:
            sheet.column_dimensions[col].hidden = True

        # Set up Grouping in Excel by levels
        sheet.sheet_properties.outlinePr.summaryBelow = True
        sheet.row_dimensions.group(2, rowcount-2, outline_level=1)
        for i in range(len(rowsToBeAdded)-1):
            if i == 0:
                sheet.row_dimensions.group(
                    2, rowsToBeAdded[i+1], outline_level=2)
            else:
                sheet.row_dimensions.group(
                    rowsToBeAdded[i]+2, rowsToBeAdded[i+1], outline_level=2)

    wb.save(filePath)


def resultExcelFormat(outPath: str, data):
    '''
    Format each sheet, and dump metrics calculated from split function onto sheets
    ---
    outpath: Directory of file after processing step 1
    data: Metrics calculated from step 1
    '''
    wb = load_workbook(outPath)

    x = 0
    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        rowcount = sheet.max_row
        colnames = []
        for cell in sheet[1]:
            colnames.append(cell.value)

        # Format excel
        border = Border(left=Side(border_style='thin', color=colors.BLACK), right=Side(border_style='thin', color=colors.BLACK),
                        top=Side(border_style='thin', color=colors.BLACK), bottom=Side(border_style='thin', color=colors.BLACK))
        leftalign = Alignment(
            horizontal='left', vertical='center', wrap_text=True)
        centeralign = Alignment(
            horizontal='center', vertical='center', wrap_text=True)
        color = Font(color='00FF0000')
        for i in range(1, rowcount+1):
            for j in range(1, len(colnames)+1):
                sheet.cell(i, j).border = border
                if sheet.cell(i, j).value == '浙江高盛输变电设备股份有限公司':
                    posx = i
                    sheet.cell(i, j).font = color
                if colnames[j-1] == '投标人名称':
                    sheet.cell(i, j).alignment = leftalign
                elif colnames[j-1] == '开标备注':
                    sheet.cell(i, j).number_format = "0.00%"
                    sheet.cell(i, j).alignment = centeralign
                else:
                    if colnames[j-1] == '投标价格':
                        posy = j
                    sheet.cell(i, j).alignment = centeralign

        # Column Width Adjustments, find column width
        for col in sheet.columns:
            curwidth = len(re.sub(
                "[A-Za-z0-9\!\%\[\]\,\。]", "", str(col[0].value))) + len(str(col[0].value))
            for j in range(len(col)):
                chinese = len(
                    re.sub("[A-Za-z0-9\!\%\[\]\,\。]", "", str(col[j].value)))
                if curwidth < chinese + len(str(col[j].value)):
                    curwidth = chinese + len(str(col[j].value))
            sheet.column_dimensions[col[0].column_letter].width = curwidth+0.8

        # Row Width
        for i in range(1, rowcount+1):
            sheet.row_dimensions[i].height = 17

        # Hide irrelevant columns
        essentials = ['No.', '采购项目名称', '采购项目编号', '分标名称', '分包名称', '项目单位',
                      '投标人名称', '投标价格', '开标备注', '得分']
        notinlist = list(set(colnames).difference(essentials))
        notindex = []
        for each in notinlist:
            notindex.append(chr(ord('@')+(colnames.index(each)+1)))
        for col in notindex:
            sheet.column_dimensions[col].hidden = True

        # document calculations onto the sheets
        try:
            sheet.cell(posx, len(colnames)+1).value = data[x][6]
            sheet.cell(posx, len(colnames)+1).number_format = "0.00%"
        except:
            pass
        sheet.cell(1, len(colnames)+2).value = "A1"
        sheet.cell(1, len(colnames)+3).value = data[x][0]
        sheet.cell(2, len(colnames)+2).value = "A1区间"
        sheet.cell(2, len(colnames)+3).value = data[x][0]*0.85
        sheet.cell(2, len(colnames)+4).value = data[x][0]*1.1
        sheet.cell(3, len(colnames)+2).value = "Check?"
        sheet.cell(3, len(colnames)+3).value = data[x][1]
        sheet.cell(3, len(colnames)+4).value = data[x][2]
        sheet.cell(4, len(colnames)+2).value = "A2"
        sheet.cell(4, len(colnames)+3).value = data[x][3]
        sheet.cell(5, len(colnames)+2).value = "C值"
        sheet.cell(5, len(colnames)+3).value = data[x][4]
        sheet.cell(5, len(colnames)+3).number_format = "0.00%"
        sheet.cell(6, len(colnames)+2).value = "基准价"
        sheet.cell(6, len(colnames)+3).value = data[x][5]
        sheet.cell(7, len(colnames)+2).value = "最佳点数"
        try:
            sheet.cell(7, len(colnames)+3).value = 1 - \
                sheet.cell(posx, posy).value*(1-data[x][6])/data[x][5]
        except:
            sheet.cell(7, len(colnames)+3).value = 0
        sheet.cell(7, len(colnames)+3).number_format = "0.00%"
        sheet.cell(8, len(colnames)+2).value = "中标点数"

        for i in range(1, 9):
            for j in range(len(colnames)+2, len(colnames)+5):
                sheet.cell(i, j).alignment = centeralign
        x = x+1
    wb.save(outPath)
