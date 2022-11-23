import pandas as pd
import numpy as np
import os
from openpyxl import *
from openpyxl.styles import Border, Side, Alignment, colors, Font
import re


def bidExcelProcessing(filePath: str, costPath: str, parameterPath: str, outPath: str) -> dict:
    """Takes Excel sheets from Sourcefile, and Processing it and output the completed file without styles
        - all path has to be full path
        - filePath: Path of input files
        - costPath: Path of Cost file
        - parameterPath: Path of Parameters file
        - outPath: Path of output file
    """
    assert ".xlsx" in filePath, f"Input FilePath given is not Excel"
    assert ".xlsx" in costPath, f"Costs FilePath given is not Excel"
    assert ".xlsx" in parameterPath, f"Parameters FilePath given is not Excel"
    assert ".xlsx" in outPath, f"Output FilePath given is not Excel"

    # Import all constants from existing files
    costFiles = pd.read_excel(costPath, sheet_name=['Sheet1', 'Sheet4'])
    parameters = pd.read_excel(parameterPath, sheet_name=None)
    bao = parameters['Sheet1']
    guige = parameters['Sheet2']
    shipping = parameters['Sheet3']

    # Main file manipulation
    project = pd.read_excel(filePath, sheet_name=None)

    passedDict = {}

    for key, subproject in project.items():
        # Check num of clients
        clients = len(set(subproject['需求单位']))

        # Identify relevant Cost File
        if '低压电力电缆' in key:
            cost = costFiles['Sheet4']
        else:
            cost = costFiles['Sheet1']

        rowcount = subproject.shape[0]
        subproject = subproject.assign(
            **{'含税单价': np.zeros(rowcount), 'CB': np.zeros(rowcount), '比例': np.zeros(rowcount)})

        for index, row in subproject.iterrows():

            # Processing step 1, rename Baos and Search relevant costs
            if len(row['包名称']) == 2:
                subproject.loc[index, '包名称'] = row['包名称'].replace("包", "包0")
            subproject.loc[index, 'CB'] = cost.loc[cost['产品名称']
                                                   == row['物资名称'], '每千米万元'].values[0]

            # Processing Step 2, Calculate Percentage based on multiple information, type, delivery method, number of clients
            bili = 0
            try:
                bili += guige.loc[(guige['项目编号'] == row['分标编号']) & (guige['项目单位'] ==
                                                                    row['项目单位']) & (guige['物资名称'] == row['物资名称']), '合计调整'].values[0]
            except:
                return "Missing 物资 in Cost File."

            if '地面' in str(row['交货地点']) or '地面' in str(row['交货方式']):
                bili += 0.01

            try:
                if clients > 1:
                    bili += shipping[shipping['需求单位'] ==
                                     row['需求单位']]['运费调整'].values[0]
            except:
                return "Missing 需求单位 in 运费调整"

            bili += bao[(bao['项目编号'] == row['分标编号']) & (bao['项目单位'] == row['项目单位'])
                        & (bao['分标编号'] == key) & (bao['包号'] == subproject.loc[index, '包名称'])]['积'].values[0]

            subproject.loc[index,
                           '含税单价'] = subproject.loc[index, 'CB'] / (1 - bili)

        subproject['CB总价'] = subproject['CB'] * subproject['数量']
        subproject['未含税单价'] = np.round(subproject['含税单价'] / 1.13, 6)
        subproject['含税总价'] = subproject['未含税单价'] * 1.13 * subproject['数量']
        subproject['比例'] = 1 - subproject['CB']/subproject['含税单价']

        new_cols = ['分标编号', '包名称', '分包编号', '项目单位', '需求单位', '项目名称', '工程电压等级', '物资名称', '物资描述',
                    '单位', '数量', '未含税单价', '含税单价', '含税总价', 'CB', 'CB总价', '比例', '首批交货日期',
                    '最后一批交货日期', '交货地点', '备注', '技术规范编码', '网省采购申请号', '总部采购申请号 ', '物料编码', '扩展描述', '扩展编码']
        subproject = subproject.reindex(columns=new_cols)
        subproject.sort_values(by=['包名称', '需求单位', '物资名称', '数量'], inplace=True)

        if os.path.isfile(outPath):
            with pd.ExcelWriter(outPath, mode="a", engine='openpyxl', if_sheet_exists="new") as writer:
                subproject.to_excel(writer, sheet_name=key, index=False)
        else:
            subproject.to_excel(outPath, sheet_name=key, index=False)


def bidExcelFormat(filePath: str):

    assert ".xlsx" in filePath, f"Input FilePath given is not Excel"

    # Using OpenPyXL to finish formating and grouping
    wb = load_workbook(filePath)

    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        rowcount = sheet.max_row
        colcount = sheet.max_column

        # Default Border, Alignment
        border = Border(left=Side(border_style='thin', color=colors.BLACK), right=Side(border_style='thin', color=colors.BLACK),
                        top=Side(border_style='thin', color=colors.BLACK), bottom=Side(border_style='thin', color=colors.BLACK))
        leftalign = Alignment(
            horizontal='left', vertical='center', wrap_text=True)
        centeralign = Alignment(
            horizontal='center', vertical='center', wrap_text=True)

        # Find the number of unique Baos
        bao = [cell.value for cell in sheet['B']]
        baoNum = len(set(bao))

        rowsToBeAdded = [2]
        i = 2

        # Set Up Subtotal Structures used in Excel
        while(i < rowcount+baoNum+1):

            if sheet.cell(i, 2).value != sheet.cell(i+1, 2).value:
                rowsToBeAdded.append(i)
                sheet.insert_rows(i+1)
                sheet.cell(i+1, 2).value = sheet.cell(i, 2).value+" Total"
                sheet.cell(i+1, 2).font = Font(bold=True)
                sheet.cell(i+1, 14).value = "=SUBTOTAL(9,N" + \
                    str(rowsToBeAdded[-2])+":N"+str(i)+")"
                sheet.cell(i+1, 16).value = "=SUBTOTAL(9,P" + \
                    str(rowsToBeAdded[-2])+":P"+str(i)+")"
                sheet.cell(i+1, 17).value = "=1-P"+str(i)+"/N"+str(i)
                i += 1
            i += 1

        # Update Rowcount and setup Grand Total
        rowcount = sheet.max_row
        sheet.cell(rowcount-1, 2).value = "Grand Total"
        sheet.cell(rowcount-1, 2).font = Font(bold=True)
        sheet.cell(rowcount-1, 14).value = "=SUBTOTAL(9, N2:N" + \
            str(rowcount-3)+")"
        sheet.cell(rowcount-1, 14).value = "=SUBTOTAL(9, P2:P" + \
            str(rowcount-3)+")"

       # Format columns using predefined rules, using col numbers instead for simplicity
        for i in range(2, rowcount):
            for j in range(1, colcount+1):
                sheet.cell(i, j).border = border
                if j in [4, 5, 8]:
                    sheet.cell(i, j).alignment = leftalign
                else:
                    if j == 17:
                        sheet.cell(i, j).number_format = "0.00%"
                    sheet.cell(i, j).alignment = centeralign

        # Column Width Adjustments, find column width
        for col in sheet.columns:
            curwidth = len(re.sub(
                "[A-Za-z0-9\!\%\[\]\,\。]", "", str(col[0].value))) + len(str(col[0].value))
            for j in range(len(col)):
                chinese = len(
                    re.sub("[A-Za-z0-9\!\%\[\]\,\。]", "", str(col[j].value)))
                if curwidth < chinese + len(str(col[j].value)):
                    curwidth = chinese + len(str(col[j].value))
            sheet.column_dimensions[col[0].column_letter].width = curwidth+0.8

        # Row Width
        for i in range(1, rowcount):
            sheet.row_dimensions[i].height = 17

        # Hide Irrelavent columns
        clients = [cell.value for cell in sheet['E']]
        invisible = ['A', 'C', 'F', 'G', 'I',
                     'R', 'U', 'V', 'X', 'Y', 'Z', 'AA']
        if len(set(clients)) > 2:
            invisible = invisible + ['D']
        else:
            invisible = invisible + ['E']
        for col in invisible:
            sheet.column_dimensions[col].hidden = True

        # Set up Grouping in Excel by levels
        sheet.sheet_properties.outlinePr.summaryBelow = True
        sheet.row_dimensions.group(2, rowcount-2, outline_level=1)
        for i in range(len(rowsToBeAdded)-1):
            if i == 0:
                sheet.row_dimensions.group(
                    2, rowsToBeAdded[i+1], outline_level=2)
            else:
                sheet.row_dimensions.group(
                    rowsToBeAdded[i]+2, rowsToBeAdded[i+1], outline_level=2)

    wb.save(filePath)


def migration(filePath, outPath):
    '''
    Mapping completed table onto worksheets ready for submission
    ---
    input: confirmed Excelsheets
    output: Excelsheets ready for submssion
    ---
    '''

    # Input and process all files, locate relevant sheets by matching number
    try:
        inputFile = pd.read_excel(
            filePath, sheet_name=None, converters={'网省采购申请号': int})
        table = pd.read_excel(
            outPath, skiprows=0, usecols='A:S', converters={'网省采购申请行号': int})
        table = table.drop('未含税单价(万)', axis=1)
    except:
        return "File format might have changed, check source files."

    if len(table.loc[0, '分包名称']) == 2:
        table.loc[:, '分包名称'] = table.loc[:, '分包名称'].apply(
            lambda x: x.replace('包', '包0') if len(x) == 2 else x)
    for _, sheet in inputFile.items():
        if sheet.loc[0, '网省采购申请号'] in list(table['网省采购申请行号']):
            inputSheet = sheet
            break

    # Migrate data
    temp = inputSheet.loc[:, ['包名称', '网省采购申请号', '物资名称', '未含税单价']]
    temp = temp.rename(columns={
                       '包名称': '分包名称', '网省采购申请号': '网省采购申请行号', '物资名称': '物料名称', '未含税单价': '未含税单价(万)'})
    try:
        table = pd.merge(table, temp, on=['分包名称', '物料名称', '网省采购申请行号'])
    except:
        return f"{outPath.split('/')[-1].split('.')[0]} 存在数据问题"

    # Followup with computations
    table['税率（%）'] = 13
    table['含税单价(万)'] = table['未含税单价(万)'] * (1 + table['税率（%）']/100)
    table['未含税合价(万)'] = table['未含税单价(万)'] * table['数量']
    table['含税合价(万)'] = table['含税单价(万)'] * table['数量']

    # Check if two totals is the same
    if not np.isclose(sheet.tail(1)['含税总价'].values[0], table.sum()['含税合价(万)'], rtol=0, atol=1e-10):
        return f"{outPath.split('/')[-1].split('.')[0]} 与表格总价差异较大"

    new_cols = ['分包编号', '分包名称', '轮次', '附件上传状态', '物料编码', '物料名称', '技术规范书ID', '网省采购申请行号',
                '项目单位', '单位', '扩展描述', '包限价(万)', '行限价(万)', '数量', '未含税单价(万)', '税率（%）', '含税单价(万)',
                '未含税合价(万)', '含税合价(万)']
    finalTable = table.reindex(columns=new_cols)

    # Safeguard against price ceilings
    if any(table['行限价(万)'].astype(str).str.contains('\d', regex=True)):
        if any(table['行限价(万)'] < table['含税合价(万)']):
            return f"{outPath.split('/')[-1].split('.')[0]} 超过行限价"

    with pd.ExcelWriter(outPath, mode='a', engine='openpyxl', if_sheet_exists='new') as writer:
        finalTable.to_excel(writer, sheet_name='报价方式-单价',
                            index=False, startrow=1, startcol=1)
