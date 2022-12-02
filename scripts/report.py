from openpyxl import *
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from components import dataframeToImg


def bidResultReporting(outPath: str, bidder: str):
    '''
    Gather information from the bidding report into strings
    While printing, gather information about statistics of performance
    ---
    outPath: Directory of the output file
    bidder: Bidders wished to tally
    '''

    assert ".xlsx" in outPath, f"Input FilePath given is not Excel"

    file = pd.read_excel(outPath, sheet_name=None)

    try:
        projectName = file[list(file.keys())[0]].loc[0, '项目单位']
    except:
        projectName = file[list(file.keys())[0]].loc[0, '采购项目名称']
    subprojectName = file[list(file.keys())[0]].loc[0, '分标名称']
    report = {}

    # Provided additional protection for bidders not participating in the subproject
    for _, sheet in file.items():
        try:
            report[sheet.loc[0, '分包名称']] = round(sheet[sheet['投标人名称'] == bidder]['投标价格'].values[0], 2), round(sheet[sheet['投标人名称'] == bidder]['得分'].values[0], 2), sheet[sheet['投标人名称']
                                                                                                                                                                         == bidder].index[0]+1, sheet['投标人名称'].count(), sorted(sheet['得分'], reverse=True).index(sheet[sheet['投标人名称'] == bidder]['得分'].values[0])+1
        except:
            continue

    if not report:
        return f"{bidder} 没有参加 {projectName} 项目投标."

    _, ax = plt.subplots(figsize=(10, max(len(report) // 2.5, 4)))
    result = pd.DataFrame.from_dict(report, orient='index', columns=[
                                    '金额', '得分', '排名', '厂家数', '名次'])
    result = result.reset_index(names="包号")
    title = "采购单位: " + projectName + '\n' + "分标名称: " + subprojectName
    thisTitle = title + '\n' + "投标人名称: " + bidder + '\n' + "平均得分: " + str(np.average([x[1] for x in list(
        report.values())])) + '\n' + "平均名次: " + str(np.average([x[-1] for x in list(report.values())]))
    dataframeToImg(ax, result, thisTitle)

    return title, result


def finalResultReporting(successBidPath: str, resultPath: str, title: str, data):
    '''
    Mapping table to original worksheets
    ---
    input: docx successful bidder tally
    xlsx original worksheet
    ---
    '''

    assert ".xlsx" in successBidPath, f"Input FilePath given is not Excel"
    assert ".xlsx" in resultPath, f"Result FilePath given is not Excel"

    # Locate Column with success bidders either from year tally or project specific
    successBidTally = pd.read_excel(successBidPath)
    succuessBider = np.where(successBidTally.columns.str.contains('中标'))[0][0]

    # Tally the location of successful bidders in each file
    resultFiles = pd.read_excel(resultPath, sheet_name=None)
    resultProjectNum = resultFiles[list(resultFiles.keys())[0]].loc[0, '分标编号']
    relevant = list(successBidTally[successBidTally['分标编号']
                    == resultProjectNum].iloc[:, succuessBider])
    winner = relevant.copy()
    successRank = []
    priceRank = []
    successColumn = []
    winnerPrice = []
    successScore = []
    for _, sheet in resultFiles.items():
        winnerPos = np.where(sheet['投标人名称'] == relevant.pop(0))[0][0]
        successRank.append(winnerPos)
        successScore.append(round(sheet.loc[winnerPos, '得分'], 2))
        priceRank.append(int(sheet['得分'].rank(
            ascending=False)[int(winnerPos)]))
        successColumn.append(np.where(sheet.columns == '得分')[0][0])
        winnerPrice.append(sheet.loc[winnerPos, '投标价格'])
        while not pd.isna(sheet.iloc[successRank[-1], successColumn[-1]+1]):
            successColumn[-1] += 1
    # print("中标得分: ", successScore)
    # print("分数排名: ", priceRank)
    # Need a data detector here, to avoid tagging multiple time
    # If "中标" exist in any of the columns, then next do the following steps

    successDf = pd.DataFrame(
        {"中标人": winner, '中标得分': successScore, "分数排名": priceRank})
    successBoard = pd.concat([data[['包号', '得分', '排名']], successDf], axis=1)

    _, ax = plt.subplots(
        figsize=(10, max(len(successBoard) // 2.5, 4)))
    thisTitle = title + '\n' + "中标结果"
    dataframeToImg(ax, successBoard, thisTitle)

    wb = load_workbook(resultPath)
    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        sheet.cell(row=successRank.pop(0)+2,
                   column=successColumn.pop(0)+2).value = '中标'
        anchorPrice = sheet.cell(row=6, column=sheet.max_column-1).value
        anchorPoint = sheet.cell(row=7, column=sheet.max_column-1).value
        sheet.cell(row=8, column=sheet.max_column-1).value = 1 - \
            anchorPrice*(1-anchorPoint)/winnerPrice.pop(0)
        sheet.cell(row=8, column=sheet.max_column-1).number_format = "0.00%"
    wb.save(resultPath)
