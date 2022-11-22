import pandas as pd
import numpy as np
from openpyxl import *
from openpyxl.styles import PatternFill, Border, Side, Alignment, colors
from openpyxl.utils import get_column_letter
import re


def filterDownloadExcel(downloadPath: str, savePath='~/Desktop/相关清单/'):
    '''
    Filter Excel files containing multiple products downloaded from ECP 2.0
    '''

    # Input Type check
    assert ".xlsx" in downloadPath, f"Downloaded FilePath given is not Excel"

    try:
        # Import Files into Pandas DataFrame
        downloadFile = pd.read_excel(downloadPath, sheet_name=None)

        # Future Expansions possible into other products
        relatedProd = ['架空绝缘导线', '集束绝缘导线', '架空线']
        prodList = [*downloadFile]

        fileName = ""
        for key in prodList:
            if not any(prod in key for prod in relatedProd):
                file = downloadFile.pop(key)
                if not fileName:
                    # Extract fileName from Owner
                    fileName = list(file['项目单位'])[0].split(
                        '国网')[1].split('电力')[0]

        if not downloadFile:
            return 'No related subprojects present in this file.'

    except:
        return "Erros occured during extraction Process."

    # Writing Found Sheets into new sheets
    try:
        saveFullPath = savePath + fileName + "相关清单.xlsx"
        with pd.ExcelWriter(saveFullPath, engine="openpyxl") as writer:
            for key in downloadFile:
                downloadFile[key].to_excel(writer, sheet_name=key, index=False)
        print(
            f"Subprojects successully filtered. {len(downloadFile)} related subprojects found.")
        return saveFullPath

    except:
        return "Errors occured during writing process."


def relatedSheetProcessing(relatedSheetPath: str, parameterPath: str):
    '''
    Collect product information, type, amount, and delivery location
    '''
    # Check input type
    assert ".xlsx" in relatedSheetPath, f"Related Sheet FilePath given is not Excel"
    assert ".xlsx" in parameterPath, f"Parameters FilePath given is not Excel"

    relatedSheet = pd.read_excel(relatedSheetPath, sheet_name=None)
    parameterSheet = pd.read_excel(parameterPath, sheet_name=None)

    for name, sheet in relatedSheet.items():

        # Collect information from relatedSheet
        projectCode = sheet.loc[1, '分标编号']
        projectOwner = sheet.loc[1, '项目单位']
        projectName = name
        locToBeAdded = []

        # Modify information in the worksheet
        projectBao = sorted([x.replace("包", "包0") if len(
            x) == 2 else x for x in list(set(sheet.loc[:, '包名称']))])
        projectType = sorted(list(set(sheet.loc[:, '物资名称'])))
        projectInfo = [[projectCode, projectOwner, projectName, bao]
                       for bao in projectBao]
        projectInfo.sort()
        projectInfoDf = pd.DataFrame(projectInfo)
        projectProd = [[projectCode, projectOwner, prod]
                       for prod in projectType]
        projectProd.sort()
        projectProdDf = pd.DataFrame(projectProd)

        # Add unknown location to the list if any
        projectDeliveryLoc = sheet.loc[:, '需求单位'].unique()
        locParameters = parameterSheet['Sheet3'].loc[:, '需求单位'].values
        if len(projectDeliveryLoc) > 1:
            for deliveryLoc in projectDeliveryLoc:
                if deliveryLoc not in locParameters:
                    locToBeAdded.append(deliveryLoc)
        locDf = pd.DataFrame(locToBeAdded)

        # Paste info to Parameters
        with pd.ExcelWriter(parameterPath, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            projectInfoDf.to_excel(writer, sheet_name="Sheet1",
                                   startrow=writer.sheets['Sheet1'].max_row, index=False, header=False)
            projectProdDf.to_excel(writer, sheet_name="Sheet2",
                                   startrow=writer.sheets['Sheet2'].max_row, index=False, header=False)
            locDf.to_excel(writer, sheet_name="Sheet3",
                           startrow=writer.sheets['Sheet3'].max_row, index=False, header=False)
